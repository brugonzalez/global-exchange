from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import TemplateView, ListView, CreateView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse, HttpResponse, Http404
from django.db.models import Sum, Count, Avg, Q
from django.utils import timezone
from django.contrib import messages
from django.urls import reverse_lazy, reverse
from datetime import timedelta, datetime
from decimal import Decimal
import json
import io
import csv
import logging

# Importaciones para generación de PDF y Excel
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import xlsxwriter

from transacciones.models import Transaccion
from divisas.models import Moneda, TasaCambio, HistorialTasaCambio
from .models import Reporte, MetricaPanel, PlantillaReporte
from .forms import FormularioGeneracionReporte
from cuentas.views import MixinPermisosAdmin

logger = logging.getLogger(__name__)


class VistaListaReportes(LoginRequiredMixin, MixinPermisosAdmin, ListView):
    """Lista todos los reportes del usuario actual"""
    model = Reporte
    template_name = 'reportes/lista_reportes.html'
    context_object_name = 'reportes'
    paginate_by = 20
    permiso_requerido = 'ver_reportes'

    def get_queryset(self):
        return Reporte.objects.filter(solicitado_por=self.request.user)


class VistaCrearReporte(LoginRequiredMixin, MixinPermisosAdmin, CreateView):
    """Crea un nuevo reporte"""
    model = Reporte
    form_class = FormularioGeneracionReporte
    template_name = 'reportes/crear_reporte.html'
    success_url = reverse_lazy('reportes:lista_reportes')
    permiso_requerido = 'ver_reportes'

    def form_valid(self, formulario):
        formulario.instance.solicitado_por = self.request.user
        respuesta = super().form_valid(formulario)
        
        # Generar el reporte de forma síncrona
        try:
            from .tasks import generar_reporte_sincrono
            generar_reporte_sincrono(self.object.id)
            messages.success(
                self.request,
                f'Reporte "{self.object.nombre_reporte}" generado exitosamente.'
            )
        except Exception as e:
            logger.error(f"Error al generar el reporte {self.object.id} (síncrono): {e}")
            messages.warning(
                self.request,
                'Reporte creado pero hubo un problema durante la generación. Contacte a soporte.'
            )
        
        return respuesta

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['usuario'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['formulario'] = context['form']
        return context


class VistaDetalleReporte(LoginRequiredMixin, MixinPermisosAdmin, DetailView):
    """Ve los detalles de un reporte"""
    model = Reporte
    template_name = 'reportes/detalle_reporte.html'
    context_object_name = 'reporte'
    permiso_requerido = 'ver_reportes'

    def get_queryset(self):
        return Reporte.objects.filter(solicitado_por=self.request.user)


class VistaDescargarReporte(LoginRequiredMixin, MixinPermisosAdmin, TemplateView):
    """Descarga un reporte generado"""
    permiso_requerido = 'ver_reportes'

    def get(self, solicitud, *args, **kwargs):
        id_reporte = kwargs.get('id_reporte')
        reporte = get_object_or_404(
            Reporte, 
            id=id_reporte, 
            solicitado_por=solicitud.user
        )

        if not reporte.puede_descargar():
            messages.error(solicitud, 'El reporte no está disponible para descarga.')
            return redirect('reportes:detalle_reporte', pk=reporte.id)

        if not reporte.ruta_archivo:
            messages.error(solicitud, 'El archivo del reporte no está disponible.')
            return redirect('reportes:detalle_reporte', pk=reporte.id)

        try:
            # Incrementar el contador de descargas
            reporte.incrementar_conteo_descargas()

            # Servir el archivo
            respuesta = HttpResponse()
            respuesta['Content-Type'] = self._obtener_tipo_contenido(reporte.formato)
            
            # Corregir la extensión de archivo para archivos de Excel
            extension_archivo = 'xlsx' if reporte.formato == 'EXCEL' else reporte.formato.lower()
            respuesta['Content-Disposition'] = f'attachment; filename="{reporte.nombre_reporte}.{extension_archivo}"'
            
            with open(reporte.ruta_archivo.path, 'rb') as f:
                respuesta.write(f.read())
            
            return respuesta

        except Exception as e:
            logger.error(f"Error al descargar el reporte {reporte.id}: {e}")
            messages.error(solicitud, 'Error al descargar el reporte.')
            return redirect('reportes:detalle_reporte', pk=reporte.id)

    def _obtener_tipo_contenido(self, tipo_formato):
        """Obtiene el tipo de contenido para diferentes formatos"""
        tipos_contenido = {
            'PDF': 'application/pdf',
            'EXCEL': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        }
        return tipos_contenido.get(tipo_formato, 'application/octet-stream')


class APIVistaGenerarReporte(LoginRequiredMixin, MixinPermisosAdmin, TemplateView):
    """Endpoint de API para la generación inmediata de reportes"""
    permiso_requerido = 'ver_reportes'

    def post(self, solicitud, *args, **kwargs):
        try:
            datos = json.loads(solicitud.body)
            tipo_reporte = datos.get('tipo_reporte')
            tipo_formato = datos.get('formato', 'PDF')
            fecha_desde = datos.get('fecha_desde')
            fecha_hasta = datos.get('fecha_hasta')
            filtros = datos.get('filtros', {})

            # Validar fechas
            try:
                fecha_desde = datetime.fromisoformat(fecha_desde.replace('Z', '+00:00'))
                fecha_hasta = datetime.fromisoformat(fecha_hasta.replace('Z', '+00:00'))
            except (ValueError, AttributeError):
                return JsonResponse({
                    'success': False,
                    'error': 'Formato de fecha inválido'
                })

            # Generar reporte inmediatamente para conjuntos de datos pequeños
            if tipo_reporte == 'HISTORIAL_TRANSACCIONES':
                transacciones = Transaccion.objects.filter(
                    cliente__usuarios=solicitud.user,
                    fecha_creacion__range=[fecha_desde, fecha_hasta]
                )
                
                # Aplicar filtros
                if filtros.get('moneda'):
                    transacciones = transacciones.filter(moneda_origen__codigo=filtros['moneda'])
                
                if filtros.get('estado'):
                    transacciones = transacciones.filter(estado=filtros['estado'])

                # Comprobar tamaño del conjunto de datos
                if transacciones.count() > 1000:
                    return JsonResponse({
                        'success': False,
                        'error': 'Conjunto de datos muy grande. Use la generación asíncrona.'
                    })

                # Generar contenido del reporte
                if tipo_formato == 'PDF':
                    contenido = self._generar_pdf_transacciones(transacciones, fecha_desde, fecha_hasta)
                    tipo_contenido = 'application/pdf'
                    nombre_archivo = f'transacciones_{fecha_desde.strftime("%Y%m%d")}_{fecha_hasta.strftime("%Y%m%d")}.pdf'
                elif tipo_formato == 'EXCEL':
                    contenido = self._generar_excel_transacciones(transacciones, fecha_desde, fecha_hasta)
                    tipo_contenido = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    nombre_archivo = f'transacciones_{fecha_desde.strftime("%Y%m%d")}_{fecha_hasta.strftime("%Y%m%d")}.xlsx'
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'Formato no soportado'
                    })

                # Devolver archivo como respuesta
                respuesta = HttpResponse(contenido, content_type=tipo_contenido)
                respuesta['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                return respuesta

            elif tipo_reporte == 'TASAS_CAMBIO':
                from divisas.models import TasaCambio, HistorialTasaCambio
                
                # Obtener datos de tasas de cambio
                tasas = TasaCambio.objects.filter(
                    fecha_creacion__range=[fecha_desde, fecha_hasta]
                ).select_related('moneda', 'moneda_base')
                
                # Aplicar filtro de moneda si se especifica
                if filtros.get('moneda'):
                    tasas = tasas.filter(moneda__codigo=filtros['moneda'])

                # Comprobar tamaño del conjunto de datos
                if tasas.count() > 1000:
                    return JsonResponse({
                        'success': False,
                        'error': 'Conjunto de datos muy grande. Use la generación asíncrona.'
                    })

                # Generar contenido del reporte
                if tipo_formato == 'PDF':
                    contenido = self._generar_pdf_tasas_cambio(tasas, fecha_desde, fecha_hasta)
                    tipo_contenido = 'application/pdf'
                    nombre_archivo = f'tasas_cambio_{fecha_desde.strftime("%Y%m%d")}_{fecha_hasta.strftime("%Y%m%d")}.pdf'
                elif tipo_formato == 'EXCEL':
                    contenido = self._generar_excel_tasas_cambio(tasas, fecha_desde, fecha_hasta)
                    tipo_contenido = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    nombre_archivo = f'tasas_cambio_{fecha_desde.strftime("%Y%m%d")}_{fecha_hasta.strftime("%Y%m%d")}.xlsx'
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'Formato no soportado'
                    })

                # Devolver archivo como respuesta
                respuesta = HttpResponse(contenido, content_type=tipo_contenido)
                respuesta['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
                return respuesta

            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Tipo de reporte no soportado para generación inmediata'
                })

        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'JSON inválido'
            })
        except Exception as e:
            logger.error(f"Error al generar reporte inmediato: {e}")
            return JsonResponse({
                'success': False,
                'error': 'Error interno del servidor'
            })

    def _generar_pdf_transacciones(self, transacciones, fecha_desde, fecha_hasta):
        """Genera un reporte en PDF para transacciones"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        estilos = getSampleStyleSheet()
        historia = []

        # Título
        estilo_titulo = ParagraphStyle(
            'TituloPersonalizado',
            parent=estilos['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        
        titulo = Paragraph(
            f"Reporte de Transacciones<br/>"
            f"<font size=12>Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}</font>",
            estilo_titulo
        )
        historia.append(titulo)
        historia.append(Spacer(1, 12))

        # Resumen
        total_transacciones = transacciones.count()
        try:
            volumen_total = transacciones.aggregate(total=Sum('monto_origen'))['total'] or Decimal('0')
        except (TypeError, ValueError) as e:
            logger.warning(f"Error al calcular volumen total: {e}")
            volumen_total = Decimal('0')
        
        datos_resumen = [
            ['Total de Transacciones:', str(total_transacciones)],
            ['Volumen Total:', f'{volumen_total:,.2f}'],
            ['Fecha de Generación:', timezone.now().strftime('%d/%m/%Y %H:%M')]
        ]
        
        tabla_resumen = Table(datos_resumen, colWidths=[3*inch, 2*inch])
        tabla_resumen.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        historia.append(tabla_resumen)
        historia.append(Spacer(1, 20))

        # Tabla de transacciones
        if transacciones.exists():
            # Encabezados de tabla
            datos = [['Fecha', 'Tipo', 'Moneda Origen', 'Cantidad', 'Moneda Destino', 'Estado']]
            
            # Datos de tabla
            for transaccion in transacciones[:100]:  # Limitar a las primeras 100 para el PDF
                try:
                    datos.append([
                        transaccion.fecha_creacion.strftime('%d/%m/%Y'),
                        transaccion.get_tipo_transaccion_display(),
                        transaccion.moneda_origen.codigo,
                        f'{transaccion.monto_origen or Decimal("0"):,.2f}',
                        transaccion.moneda_destino.codigo,
                        transaccion.get_estado_display()
                    ])
                except (TypeError, ValueError, AttributeError) as e:
                    logger.warning(f"Error al procesar transacción ID {getattr(transaccion, 'id', 'unknown')}: {e}")
                    # Agregar fila con valores por defecto en caso de error
                    datos.append([
                        getattr(transaccion, 'fecha_creacion', timezone.now()).strftime('%d/%m/%Y'),
                        'Error',
                        'N/A',
                        '0,00',
                        'N/A',
                        'Error'
                    ])
            
            # Crear tabla
            tabla = Table(datos, colWidths=[1*inch, 1*inch, 1*inch, 1.2*inch, 1*inch, 1*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            historia.append(tabla)
            
            if transacciones.count() > 100:
                nota = Paragraph(
                    f"<i>Nota: Se muestran solo las primeras 100 transacciones de {transacciones.count()} en total.</i>",
                    estilos['Normal']
                )
                historia.append(Spacer(1, 12))
                historia.append(nota)
        else:
            sin_datos = Paragraph("No se encontraron transacciones en el período seleccionado.", estilos['Normal'])
            historia.append(sin_datos)

        # Construir PDF
        doc.build(historia)
        buffer.seek(0)
        return buffer.getvalue()

    def _generar_excel_transacciones(self, transacciones, fecha_desde, fecha_hasta):
        """Genera un reporte en Excel para transacciones"""
        salida = io.BytesIO()
        libro = xlsxwriter.Workbook(salida, {'in_memory': True})
        hoja = libro.add_worksheet('Transacciones')

        # Formatos
        formato_titulo = libro.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        formato_encabezado = libro.add_format({
            'bold': True,
            'bg_color': '#366092',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        formato_celda = libro.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        formato_numero = libro.add_format({
            'num_format': '#,##0.00',
            'align': 'right',
            'valign': 'vcenter',
            'border': 1
        })

        # Título
        hoja.merge_range('A1:G2', 
            f'Reporte de Transacciones - Del {fecha_desde.strftime("%d/%m/%Y")} al {fecha_hasta.strftime("%d/%m/%Y")}',
            formato_titulo)

        # Resumen
        fila = 4
        hoja.write(fila, 0, 'Total de Transacciones:', formato_encabezado)
        hoja.write(fila, 1, transacciones.count(), formato_celda)
        
        fila += 1
        try:
            volumen_total = transacciones.aggregate(total=Sum('monto_origen'))['total'] or Decimal('0')
        except (TypeError, ValueError) as e:
            logger.warning(f"Error al calcular volumen total en Excel: {e}")
            volumen_total = Decimal('0')
        
        hoja.write(fila, 0, 'Volumen Total:', formato_encabezado)
        hoja.write(fila, 1, float(volumen_total), formato_numero)
        
        fila += 1
        hoja.write(fila, 0, 'Fecha de Generación:', formato_encabezado)
        hoja.write(fila, 1, timezone.now().strftime('%d/%m/%Y %H:%M'), formato_celda)

        # Encabezados
        fila += 3
        encabezados = ['Fecha', 'Tipo', 'Moneda Origen', 'Cantidad', 'Moneda Destino', 'Cantidad Destino', 'Estado']
        for col, encabezado in enumerate(encabezados):
            hoja.write(fila, col, encabezado, formato_encabezado)

        # Datos
        fila += 1
        for transaccion in transacciones:
            try:
                hoja.write(fila, 0, transaccion.fecha_creacion.strftime('%d/%m/%Y %H:%M'), formato_celda)
                hoja.write(fila, 1, transaccion.get_tipo_transaccion_display(), formato_celda)
                hoja.write(fila, 2, transaccion.moneda_origen.codigo, formato_celda)
                hoja.write(fila, 3, float(transaccion.monto_origen or Decimal('0')), formato_numero)
                hoja.write(fila, 4, transaccion.moneda_destino.codigo, formato_celda)
                hoja.write(fila, 5, float(transaccion.monto_destino or Decimal('0')), formato_numero)
                hoja.write(fila, 6, transaccion.get_estado_display(), formato_celda)
                fila += 1
            except (TypeError, ValueError, AttributeError) as e:
                logger.warning(f"Error al procesar transacción ID {getattr(transaccion, 'id', 'unknown')} en Excel: {e}")
                # Escribir fila con valores por defecto en caso de error
                try:
                    hoja.write(fila, 0, getattr(transaccion, 'fecha_creacion', timezone.now()).strftime('%d/%m/%Y %H:%M'), formato_celda)
                    hoja.write(fila, 1, 'Error', formato_celda)
                    hoja.write(fila, 2, 'N/A', formato_celda)
                    hoja.write(fila, 3, 0, formato_numero)
                    hoja.write(fila, 4, 'N/A', formato_celda)
                    hoja.write(fila, 5, 0, formato_numero)
                    hoja.write(fila, 6, 'Error', formato_celda)
                    fila += 1
                except Exception as inner_e:
                    logger.error(f"Error crítico al escribir fila de respaldo para transacción ID {getattr(transaccion, 'id', 'unknown')}: {inner_e}")
                    # Continuar con la siguiente transacción

        # Autoajustar anchos de columna
        for col in range(len(encabezados)):
            hoja.set_column(col, col, 15)

        libro.close()
        salida.seek(0)
        return salida.getvalue()

    def _generar_pdf_tasas_cambio(self, tasas, fecha_desde, fecha_hasta):
        """Genera un reporte en PDF para tasas de cambio"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        estilos = getSampleStyleSheet()
        historia = []

        # Título
        estilo_titulo = ParagraphStyle(
            'TituloPersonalizado',
            parent=estilos['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        
        titulo = Paragraph(
            f"Reporte de Tasas de Cambio<br/>"
            f"<font size=12>Del {fecha_desde.strftime('%d/%m/%Y')} al {fecha_hasta.strftime('%d/%m/%Y')}</font>",
            estilo_titulo
        )
        historia.append(titulo)
        historia.append(Spacer(1, 12))

        # Resumen
        total_tasas = tasas.count()
        monedas_unicas = tasas.values('moneda__codigo').distinct().count()
        
        datos_resumen = [
            ['Total de Registros:', str(total_tasas)],
            ['Monedas Diferentes:', str(monedas_unicas)],
            ['Fecha de Generación:', timezone.now().strftime('%d/%m/%Y %H:%M')]
        ]
        
        tabla_resumen = Table(datos_resumen, colWidths=[3*inch, 2*inch])
        tabla_resumen.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        historia.append(tabla_resumen)
        historia.append(Spacer(1, 20))

        # Tabla de tasas de cambio
        if tasas.exists():
            # Encabezados de tabla
            datos = [['Fecha', 'Moneda Origen', 'Moneda Destino', 'Tasa Compra', 'Tasa Venta', 'Diferencial']]
            
            # Datos de tabla
            for tasa in tasas[:100]:  # Limitar a los primeros 100 para el PDF
                try:
                    # Manejo seguro de operaciones decimales
                    if tasa.tasa_venta is not None and tasa.tasa_compra is not None:
                        diferencial = tasa.tasa_venta - tasa.tasa_compra
                    else:
                        diferencial = Decimal('0')
                    
                    datos.append([
                        tasa.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
                        tasa.moneda.codigo,
                        tasa.moneda_base.codigo,
                        f'{tasa.tasa_compra or Decimal("0"):,.4f}',
                        f'{tasa.tasa_venta or Decimal("0"):,.4f}',
                        f'{diferencial:,.4f}'
                    ])
                except (TypeError, ValueError, AttributeError) as e:
                    logger.warning(f"Error al procesar tasa de cambio ID {tasa.id}: {e}")
                    # Agregar fila con valores por defecto en caso de error
                    datos.append([
                        tasa.fecha_creacion.strftime('%d/%m/%Y %H:%M') if hasattr(tasa, 'fecha_creacion') else 'N/A',
                        getattr(tasa.moneda, 'codigo', 'N/A') if hasattr(tasa, 'moneda') else 'N/A',
                        getattr(tasa.moneda_base, 'codigo', 'N/A') if hasattr(tasa, 'moneda_base') else 'N/A',
                        'N/A',
                        'N/A',
                        'N/A'
                    ])
            
            # Crear tabla
            tabla = Table(datos, colWidths=[1.2*inch, 1*inch, 1*inch, 1.2*inch, 1.2*inch, 1*inch])
            tabla.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            historia.append(tabla)
            
            if tasas.count() > 100:
                nota = Paragraph(
                    f"<i>Nota: Se muestran solo los primeros 100 registros de {tasas.count()} en total.</i>",
                    estilos['Normal']
                )
                historia.append(Spacer(1, 12))
                historia.append(nota)
        else:
            sin_datos = Paragraph("No se encontraron tasas de cambio en el período seleccionado.", estilos['Normal'])
            historia.append(sin_datos)

        # Construir PDF
        doc.build(historia)
        buffer.seek(0)
        return buffer.getvalue()

    def _generar_excel_tasas_cambio(self, tasas, fecha_desde, fecha_hasta):
        """Genera un reporte en Excel para tasas de cambio"""
        salida = io.BytesIO()
        libro = xlsxwriter.Workbook(salida, {'in_memory': True})
        hoja = libro.add_worksheet('Tasas de Cambio')

        # Formatos
        formato_titulo = libro.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'vcenter'
        })
        
        formato_encabezado = libro.add_format({
            'bold': True,
            'bg_color': '#366092',
            'font_color': 'white',
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        formato_celda = libro.add_format({
            'align': 'center',
            'valign': 'vcenter',
            'border': 1
        })
        
        formato_numero = libro.add_format({
            'num_format': '#,##0.0000',
            'align': 'right',
            'valign': 'vcenter',
            'border': 1
        })

        # Título
        hoja.merge_range('A1:G2', 
            f'Reporte de Tasas de Cambio - Del {fecha_desde.strftime("%d/%m/%Y")} al {fecha_hasta.strftime("%d/%m/%Y")}',
            formato_titulo)

        # Resumen
        fila = 4
        hoja.write(fila, 0, 'Total de Registros:', formato_encabezado)
        hoja.write(fila, 1, tasas.count(), formato_celda)
        
        fila += 1
        monedas_unicas = tasas.values('moneda__codigo').distinct().count()
        hoja.write(fila, 0, 'Monedas Diferentes:', formato_encabezado)
        hoja.write(fila, 1, monedas_unicas, formato_celda)
        
        fila += 1
        hoja.write(fila, 0, 'Fecha de Generación:', formato_encabezado)
        hoja.write(fila, 1, timezone.now().strftime('%d/%m/%Y %H:%M'), formato_celda)

        # Encabezados
        fila += 3
        encabezados = ['Fecha', 'Moneda Origen', 'Moneda Destino', 'Tasa Compra', 'Tasa Venta', 'Diferencial', 'Estado']
        for col, encabezado in enumerate(encabezados):
            hoja.write(fila, col, encabezado, formato_encabezado)

        # Datos
        fila += 1
        for tasa in tasas:
            try:
                # Manejo seguro de operaciones decimales
                if tasa.tasa_venta is not None and tasa.tasa_compra is not None:
                    diferencial = tasa.tasa_venta - tasa.tasa_compra
                else:
                    diferencial = Decimal('0')
                
                hoja.write(fila, 0, tasa.fecha_creacion.strftime('%d/%m/%Y %H:%M'), formato_celda)
                hoja.write(fila, 1, tasa.moneda.codigo, formato_celda)
                hoja.write(fila, 2, tasa.moneda_base.codigo, formato_celda)
                hoja.write(fila, 3, float(tasa.tasa_compra or Decimal('0')), formato_numero)
                hoja.write(fila, 4, float(tasa.tasa_venta or Decimal('0')), formato_numero)
                hoja.write(fila, 5, float(diferencial), formato_numero)
                hoja.write(fila, 6, 'Activa' if tasa.esta_activa else 'Inactiva', formato_celda)
                fila += 1
            except (TypeError, ValueError, AttributeError) as e:
                logger.warning(f"Error al procesar tasa de cambio ID {tasa.id} en Excel: {e}")
                # Escribir fila con valores por defecto en caso de error
                try:
                    hoja.write(fila, 0, tasa.fecha_creacion.strftime('%d/%m/%Y %H:%M') if hasattr(tasa, 'fecha_creacion') else 'N/A', formato_celda)
                    hoja.write(fila, 1, getattr(tasa.moneda, 'codigo', 'N/A') if hasattr(tasa, 'moneda') else 'N/A', formato_celda)
                    hoja.write(fila, 2, getattr(tasa.moneda_base, 'codigo', 'N/A') if hasattr(tasa, 'moneda_base') else 'N/A', formato_celda)
                    hoja.write(fila, 3, 0, formato_numero)
                    hoja.write(fila, 4, 0, formato_numero)
                    hoja.write(fila, 5, 0, formato_numero)
                    hoja.write(fila, 6, 'Error', formato_celda)
                    fila += 1
                except Exception as inner_e:
                    logger.error(f"Error crítico al escribir fila de respaldo para tasa ID {getattr(tasa, 'id', 'unknown')}: {inner_e}")
                    # Continuar con la siguiente tasa

        # Autoajustar anchos de columna
        for col in range(len(encabezados)):
            hoja.set_column(col, col, 15)

        libro.close()
        salida.seek(0)
        return salida.getvalue()


class VistaPanelAnaliticas(LoginRequiredMixin, MixinPermisosAdmin, TemplateView):
    """
    Panel de analíticas principal con métricas de negocio.
    """
    template_name = 'reportes/panel_analiticas.html'
    permiso_requerido = 'ver_reportes'
    
    def get_context_data(self, **kwargs):
        contexto = super().get_context_data(**kwargs)
        
        # Obtener períodos de tiempo para el análisis
        ahora = timezone.now()
        inicio_hoy = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
        inicio_semana = inicio_hoy - timedelta(days=7)
        inicio_mes = inicio_hoy - timedelta(days=30)
        
        # Obtener datos de transacciones
        transacciones_hoy = Transaccion.objects.filter(fecha_creacion__gte=inicio_hoy)
        transacciones_semana = Transaccion.objects.filter(fecha_creacion__gte=inicio_semana)
        transacciones_mes = Transaccion.objects.filter(fecha_creacion__gte=inicio_mes)
        
        # Calcular métricas básicas
        contexto.update({
            'conteo_transacciones_hoy': transacciones_hoy.count(),
            'conteo_transacciones_semana': transacciones_semana.count(),
            'conteo_transacciones_mes': transacciones_mes.count(),
            
            'volumen_hoy': transacciones_hoy.aggregate(
                total=Sum('monto_origen')
            )['total'] or Decimal('0'),
            
            'volumen_semana': transacciones_semana.aggregate(
                total=Sum('monto_origen')
            )['total'] or Decimal('0'),
            
            'volumen_mes': transacciones_mes.aggregate(
                total=Sum('monto_origen')
            )['total'] or Decimal('0'),
            
            'total_monedas_activas': Moneda.objects.filter(esta_activa=True).count(),
            'total_clientes': Transaccion.objects.values('cliente').distinct().count(),
        })
        
        return contexto


class VistaAnalisisGanancias(LoginRequiredMixin, MixinPermisosAdmin, TemplateView):
    """
    Análisis detallado y monitoreo de ganancias.
    """
    template_name = 'reportes/analisis_ganancias.html'
    permiso_requerido = 'ver_reportes'
    
    def get_context_data(self, **kwargs):
        contexto = super().get_context_data(**kwargs)
        
        # Obtener períodos de tiempo
        ahora = timezone.now()
        inicio_hoy = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
        inicio_semana = inicio_hoy - timedelta(days=7)
        inicio_mes = inicio_hoy - timedelta(days=30)
        
        # Calcular ganancias por período
        ganancias = self.calcular_ganancias()
        
        contexto.update({
            'ganancia_hoy': ganancias['hoy'],
            'ganancia_semana': ganancias['semana'],
            'ganancia_mes': ganancias['mes'],
            'ganancia_total': ganancias['total'],
            
            'ganancia_por_moneda': ganancias['por_moneda'],
            'ganancia_por_tipo_transaccion': ganancias['por_tipo'],
            
            'monedas_principales': self.obtener_monedas_principales(),
            'tendencia_ganancias': self.obtener_tendencia_ganancias(),
        })
        
        return contexto
    
    def calcular_ganancias(self):
        """Calcula las ganancias de las transacciones completadas"""
        ahora = timezone.now()
        inicio_hoy = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
        inicio_semana = inicio_hoy - timedelta(days=7)
        inicio_mes = inicio_hoy - timedelta(days=30)
        
        # Obtener transacciones completadas (asumiendo que monto_comision representa la ganancia)
        transacciones_completadas = Transaccion.objects.filter(
            estado__in=['COMPLETADA', 'PAGADA']
        )
        
        ganancia_hoy = transacciones_completadas.filter(
            fecha_creacion__gte=inicio_hoy
        ).aggregate(total=Sum('monto_comision'))['total'] or Decimal('0')
        
        ganancia_semana = transacciones_completadas.filter(
            fecha_creacion__gte=inicio_semana
        ).aggregate(total=Sum('monto_comision'))['total'] or Decimal('0')
        
        ganancia_mes = transacciones_completadas.filter(
            fecha_creacion__gte=inicio_mes
        ).aggregate(total=Sum('monto_comision'))['total'] or Decimal('0')
        
        ganancia_total = transacciones_completadas.aggregate(
            total=Sum('monto_comision')
        )['total'] or Decimal('0')
        
        # Ganancia por moneda
        ganancia_por_moneda = []
        monedas = Moneda.objects.filter(esta_activa=True)
        for moneda in monedas:
            ganancia_moneda = transacciones_completadas.filter(
                moneda_origen=moneda
            ).aggregate(total=Sum('monto_comision'))['total'] or Decimal('0')
            
            if ganancia_moneda > 0:
                ganancia_por_moneda.append({
                    'moneda': moneda,
                    'ganancia': ganancia_moneda
                })
        
        # Ganancia por tipo de transacción
        ganancia_compra = transacciones_completadas.filter(
            tipo_transaccion='COMPRA'
        ).aggregate(total=Sum('monto_comision'))['total'] or Decimal('0')
        
        ganancia_venta = transacciones_completadas.filter(
            tipo_transaccion='VENTA'
        ).aggregate(total=Sum('monto_comision'))['total'] or Decimal('0')
        
        return {
            'hoy': ganancia_hoy,
            'semana': ganancia_semana,
            'mes': ganancia_mes,
            'total': ganancia_total,
            'por_moneda': ganancia_por_moneda,
            'por_tipo': {
                'compra': ganancia_compra,
                'venta': ganancia_venta
            }
        }
    
    def obtener_monedas_principales(self):
        """Obtiene las monedas principales por volumen de transacción"""
        return Transaccion.objects.filter(
            estado__in=['COMPLETADA', 'PAGADA']
        ).values(
            'moneda_origen__codigo',
            'moneda_origen__nombre',
            'moneda_origen__simbolo'
        ).annotate(
            conteo_transacciones=Count('id'),
            volumen_total=Sum('monto_origen')
        ).order_by('-volumen_total')[:5]
    
    def obtener_tendencia_ganancias(self):
        """Obtiene la tendencia de ganancias de los últimos 7 días"""
        hoy = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
        datos_tendencia = []
        
        for i in range(7):
            dia = hoy - timedelta(days=i)
            fin_dia = dia + timedelta(days=1)
            
            ganancia_dia = Transaccion.objects.filter(
                estado__in=['COMPLETADA', 'PAGADA'],
                fecha_creacion__gte=dia,
                fecha_creacion__lt=fin_dia
            ).aggregate(total=Sum('monto_comision'))['total'] or Decimal('0')
            
            datos_tendencia.append({
                'fecha': dia.strftime('%Y-%m-%d'),
                'ganancia': float(ganancia_dia)
            })
        
        return list(reversed(datos_tendencia))


class APIVistaMetricasPanel(LoginRequiredMixin, MixinPermisosAdmin, TemplateView):
    """
    Endpoint de API para métricas del panel en tiempo real.
    """
    permiso_requerido = 'ver_reportes'
    
    def get(self, solicitud, *args, **kwargs):
        try:
            # Obtener métricas actuales
            ahora = timezone.now()
            inicio_hoy = ahora.replace(hour=0, minute=0, second=0, microsecond=0)
            
            # Métricas de transacción
            transacciones_hoy = Transaccion.objects.filter(fecha_creacion__gte=inicio_hoy)
            completadas_hoy = transacciones_hoy.filter(estado__in=['COMPLETADA', 'PAGADA'])
            
            # Cálculo de ganancias
            ganancia_hoy = completadas_hoy.aggregate(
                total=Sum('monto_comision')
            )['total'] or Decimal('0')
            
            # Cálculo de volumen
            volumen_hoy = transacciones_hoy.aggregate(
                total=Sum('monto_origen')
            )['total'] or Decimal('0')
            
            # Estimación de ganancia por diferencial de tasa de cambio
            ganancia_total_diferencial = Decimal('0')
            for moneda in Moneda.objects.filter(esta_activa=True):
                tasa = moneda.obtener_tasa_actual()
                if tasa:
                    diferencial = tasa.tasa_venta - tasa.tasa_compra
                    volumen_moneda = transacciones_hoy.filter(
                        moneda_origen=moneda
                    ).aggregate(total=Sum('monto_origen'))['total'] or Decimal('0')
                    
                    ganancia_diferencial = diferencial * volumen_moneda * Decimal('0.1')  # Estimar 10% del diferencial
                    ganancia_total_diferencial += ganancia_diferencial
            
            return JsonResponse({
                'success': True,
                'metricas': {
                    'transacciones_hoy': transacciones_hoy.count(),
                    'ganancia_hoy': float(ganancia_hoy),
                    'volumen_hoy': float(volumen_hoy),
                    'ganancia_estimada_diferencial': float(ganancia_total_diferencial),
                    'estimacion_ganancia_total': float(ganancia_hoy + ganancia_total_diferencial),
                    'monedas_activas': Moneda.objects.filter(esta_activa=True).count(),
                    'transacciones_pendientes': Transaccion.objects.filter(
                        estado='PENDIENTE'
                    ).count(),
                },
                'timestamp': ahora.isoformat()
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })